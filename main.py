from openpyxl import Workbook
from openpyxl import load_workbook
import os

def create_csv(file_name: str) -> None:
    wb = load_workbook(file_name, data_only=True)
    ws = wb.active

    with open(f"Listas/{file_name[file_name.find('/'):file_name.find('.')]}.csv", "w", encoding='ansi') as file:
        file.write("remark;CSV-Import-Export;;;;;\nremark;Date = Wed Sep 18 05:04:03 2024;;;;;\nremark;Version = RSLogix 5000 v36.00;;;;;\nremark;Owner = ;;;;;\nremark;Company = ;;;;;\n0.3;;;;;;\nTYPE;SCOPE;NAME;DESCRIPTION;DATATYPE;SPECIFIER;ATTRIBUTES\n")
        #linha_inicial = int(input("Aonde começam os dados da planilha (onde acabam cabeçalhos)?"))
        linha_inicial = 8
        colunas = (8, 11)

        linha = []
        
        for row in ws.iter_rows(min_row=linha_inicial, min_col=colunas[0], max_col=colunas[-1]):
            for cell in row:
                linha.append(cell.value)
            linha.pop(2)
            if linha[0] != None or linha[1] != None or linha[2] != None:
                file.write(f"ALIAS;;{linha[0]};{linha[1]};;{linha[2]};(RADIX := Decimal, ExternalAccess := Read/Write)\n")
            linha.clear()

def main() -> None:
    for _path, _subdirs, files in os.walk((os.getcwd()).replace('\\', '/') + '/PTS/'):
        for file in files:
            if file.endswith('.xlsx') or file.endswith('.xls') or file.endswith('.XLS'):
                create_csv(f"PTS/{file}")


if __name__ == "__main__":
    main()